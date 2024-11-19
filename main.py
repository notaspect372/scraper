import requests
from bs4 import BeautifulSoup
import pandas as pd
import math
import re
from urllib.parse import urlparse
from openpyxl import load_workbook
from openpyxl.styles import Font
from geopy.geocoders import Nominatim
import time
import os

# Headers to mimic browser requests
headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.36'}

# Initialize geolocator
geolocator = Nominatim(user_agent="property_scraper")

start_urls = [
    "https://www.bproperty.com/rent/residential/?price=20000-25000"
]

# Function to perform requests with manual retry logic
def get_with_retry(url, retries=5, delay=2):
    attempt = 0
    while attempt < retries:
        try:
            response = requests.get(url, headers=headers, timeout=10)
            if response.status_code == 200:
                return response
            else:
                print(f"Failed with status code {response.status_code}. Retrying...")
        except requests.exceptions.RequestException as e:
            print(f"Error fetching {url}: {e}. Retrying...")
        attempt += 1
        time.sleep(delay)  # Delay between retries
    return None

def get_total_pages(url):
    try:
        response = get_with_retry(url)
        print(response.status_code)
        if response is None:
            return 1
        soup = BeautifulSoup(response.content, 'html.parser')
        total_listing_selector = 'span.CountTitle-number'
        total_listing_text = soup.select_one(total_listing_selector).get_text(strip=True).split()[0]
        print(total_listing_text)
        total_listing = int(total_listing_text.replace(',', ''))
        page_listing = 30  # Assuming 30 listings per page
        total_pages = math.ceil(total_listing / page_listing)
        return total_pages
    except Exception as e:
        print(f"Error fetching total pages for {url}: {e}")
        return 1

def get_geolocation(address):
    try:
        location = geolocator.geocode(address, timeout=10)
        if location:
            return location.longitude, location.latitude
        return None, None
    except Exception as e:
        print(f"Error getting geolocation for address {address}: {e}")
        return None, None

def parse_property(url, property_type):
    try:
        response = get_with_retry(url)
        if response is None:
            return None
        soup = BeautifulSoup(response.content, 'html.parser')
        item = {}

        name_selector = 'h1.Title-pdp-title span'
        price_selector = 'span.FirstPrice'
        description_selector = 'div.ViewMore-text-description'

        item['name'] = soup.select_one(name_selector).get_text(strip=True) if soup.select_one(name_selector) else ''
        address = soup.find('p', class_='Location')
        item['address'] = address.text.strip() if address else 'N/A'
        item['price'] = soup.select_one(price_selector).get_text(strip=True) if soup.select_one(price_selector) else ''
        item['description'] = soup.select_one(description_selector).get_text(strip=True) if soup.select_one(description_selector) else ''

        characteristics = {}
        characteristics_container = soup.find('div', class_='listing-section listing-details')
        if characteristics_container:
            details_rows = characteristics_container.find_all('div', class_='columns-2')
            for row in details_rows:
                label_div = row.find('div', class_='listing-details-label')
                value_div = row.find('div', class_='last')
                if label_div and value_div:
                    label = label_div.get_text(strip=True).lower().replace(' ', '_')
                    value = value_div.get_text(strip=True)
                    characteristics[label] = value

        item['characteristics'] = characteristics
        item['area'] = characteristics.get('floor_area(sqft)', 'N/A')

        # Transaction type
        if 'buy' in url:
            item['transaction_type'] = 'for sale'
        elif 'rent' in url:
            item['transaction_type'] = 'for rent'
        else:
            item['transaction_type'] = 'unknown'

        # Use provided property type
        item['property_type'] = property_type
        item['property_url'] = url

        address = item.get('address', '')
        longitude, latitude = get_geolocation(address)
        item['longitude'] = longitude
        item['latitude'] = latitude


        amenities = []
        amenities_container = soup.find('div', class_='listing-amenities-list')
        if amenities_container:
            amenity_items = amenities_container.find_all('div', class_='listing-amenities-list-item')
            for amenity in amenity_items:
                name_span = amenity.find('span', class_='listing-amenities-name')
                if name_span:
                    amenities.append(name_span.get_text(strip=True))
        item['amenities'] = amenities

        print(item)
        return item
    except Exception as e:
        print(f"Error parsing property data from {url}: {e}")
        return None

def parse_page(url, property_type, items, seen_urls):
    try:
        response = get_with_retry(url)
        if response is None:
            return
        soup = BeautifulSoup(response.content, 'html.parser')
        property_url_selector = 'a.ListingCell-ListingLink.js-listing-link'
        property_urls = {a['href'] for a in soup.select(property_url_selector)}  # Using set comprehension
        for property_url in property_urls:
            if property_url not in seen_urls:  # Check if URL is already in the seen set
                seen_urls.add(property_url)
                item = parse_property(property_url, property_type)
                print(item)
                if item:
                    items.append(item)
        time.sleep(2)  # Adding delay to mimic human interaction
    except Exception as e:
        print(f"Error parsing page {url}: {e}")

def save_to_excel(base_url, items):
    os.makedirs('artifacts', exist_ok=True)  # Ensure 'artifacts' directory exists
    parsed_url = urlparse(base_url)
    safe_filename = re.sub(r'\W+', '_', parsed_url.netloc + parsed_url.path) + '.xlsx'
    file_path = os.path.join('artifacts', safe_filename)
    df = pd.DataFrame(items)
    df.to_excel(file_path, index=False)

    # Enhance Excel formatting
    wb = load_workbook(file_path)
    ws = wb.active
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = adjusted_width
    wb.save(file_path)


def determine_property_type(url):
    if 'residential' in url or 'commercial' in url:
        return 'residential' if 'residential' in url else 'commercial'
    return 'unknown'

def construct_pagination_url(base_url, page):
    """
    Constructs the pagination URL based on whether the base URL contains a '?'.
    """
    if '?' in base_url:
        return f"{base_url}&page={page}"
    else:
        return f"{base_url}?page={page}"

def main():
    for url in start_urls:
        print(url)
        property_type = determine_property_type(url)
        total_pages = get_total_pages(url)
        items = []
        seen_urls = set()  # Initialize a set for tracking seen URLs
        
        for page in range(1, total_pages + 1):
            paginated_url = construct_pagination_url(url, page)
            parse_page(paginated_url, property_type, items, seen_urls)
            time.sleep(1)  # To prevent rate limiting
        
        save_to_excel(url, items)

if __name__ == "__main__":
    main()

